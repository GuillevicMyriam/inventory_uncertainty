# -*- coding: utf-8 -*-
"""
Copyright Swiss Federal Office for the Environment FOEN, 2021 - 2023.

This file is part of: inventory_uncertainty_UNFCCC_CLRTAP.

inventory_uncertainty_UNFCCC_CLRTAP is a free software: 
you can redistribute it and/or modify
it under the terms of the BSD 3-Clause "New" or "Revised" License.

inventory_uncertainty_UNFCCC_CLRTAP is distributed 
in the hope that it will be useful, but WITHOUT ANY WARRANTY; 
without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
See the BSD 3-Clause "New" or "Revised" License for more details.

Created on Thu Feb  9 11:09:33 2023

"""
#import from general libraries
import openpyxl #version 2.4.7
import numpy as np
import pandas as pd
from numbers import Number

#import from local files and libraries
import utils_constant as const






#HINT Various functions hereafter are used to write results to Excel files.

def write_pr_mc_results(
        df_in: pd.DataFrame,
        df_pr: pd.DataFrame,
        df_pr_AD_EF: pd.DataFrame,
        df_mc: pd.DataFrame,
        df_mc_AD_EF: pd.DataFrame,
        index_pr_total: int,
        index_pr_proc_sector_total,
        index_pr_comp_total,
        index_pr_inv_with_without_lulucf,
        index_mc_total: int,
        index_mc_proc_sector_total,
        index_mc_comp_total,
        index_mc_inv_with_without_lulucf,
        BY_string: str,
        RY_string: str,
        no_mc: int,
        routine: int,
        filename_out,
        ) -> None:
    #HINT Write results of the uncertainty analysis to Excel for the UNECE/UNFCCC reporting
    """Write results of the uncertainty analysis to Excel for the UNECE/UNFCCC reporting.
    
    The produced Excel tables can be directly copied/pasted to a Word document.
    One Excel document is produced with several tabs:
        - tab with results of uncertainty propagation (approach 1)
        - tab with results of Monte Carlo simlations (approach 2)
        - tab with results for both approaches, for each sector
        - tab with results for both approaches, for each compound (for NID only)
        - tab with results for inventory with and without LULUCF (for NID only)
        - tab with detailed results for sector 3 Agriculture, source categories 3A, 3B, 3D (for NID only)
        - tab with detailed results for sector 4 LULUCF (for NID only)
    
    Args:
        df_in: pandas DataFrame with input uncertainty data
        df_pr: pandas DataFrame with produced uncertainties
            based on uncertainty propagation
        df_pr_AD_EF: pandas DataFrame with produced uncertainties
            for activity data and emission factors based on uncertainty propagation
        df_mc: pandas DataFrame with produced uncertainties
            for emissions based on Monte Carlo simulations
        df_mc_AD_EF: pandas DataFrame with produced uncertainties 
            for activity data and emission factors based on Monte Carlo simulations
        index_pr_total: index in df_pr where the data corresponds to the inventory total.
        index_pr_proc_sector_total: index(es) in df_pr where the data corresponds to the sums per sector.
        index_pr_comp_total: index(es) in df_pr where the data corresponds to the sums per compound (for routine = NID only).
        index_pr_inv_with_without_lulucf: index(es) in df_pr where the data corresponds to the sums
            with and without the LULUCF sector (for NID only)
        index_mc_total: index in df_mc where the data corresponds to the inventory total.
        index_mc_proc_sector_total: index(es) in df_mc where the data corresponds to the sums per sector.
        index_mc_comp_total: index(es) in df_mc where the data corresponds to the sums per compound (for routine = NID only).
        index_mc_inv_with_without_lulucf: index(es) in df_mc where the data corresponds to the sums
            with and without the LULUCF sector (for NID only)
        BY_string: string with the base year, format YYYY.
        RY_string: string with the reporting year, format YYYY.
        no_mc: number of Monte Carlo simulations.
        routine: computation routine, i.e. for NID (greenhouse gases) or IIR (pollutants).
        filename_out: name of excel file where the tabs are written and saved.       
        
    Returns:
        Excel file.
        
    Raises:
        None.
    
    """
    #no_nomenc_in = len(df)
    index_mc_output = df_mc.index[df_mc["report"] == True].tolist()  
    index_pr_output = df_pr.index[df_pr["report"] == True].tolist()

    #warning! excel tab name not more than 31 characters!
    if routine == const.ROUTINE_NID:               
        text_LULUCF_and_indirect = ", including LULUCF categories and indirect CO2 emissions"
        dEM_text = "d.EM stands for direct emission and indicates that input uncertainties are given for the emission but neither for AD nor for EF."
        nomenc_text = "IPCC category; fuel/source"
        comp_text = "Gas"
        unit_string = const.NID_UNIT_STRING
        comp_label = "GHGs"
        tab_name = comp_label
        width_narrow_col = 7.5

    elif routine == const.ROUTINE_IIR:
        text_LULUCF_and_indirect = ""
        dEM_text = ""
        nomenc_text = "NFR"
        comp_text = "Pollutant"
        unit_string = const.IIR_UNIT_STRING
        comp_label = df_mc["comp_name"][0]
        tab_name = df_mc["comp_id"][0]
        width_narrow_col = 7.5

    
    wb = openpyxl.Workbook()

    ws = wb.active #active sheet  
    ws.title = "readme"

    #===========================================================================
    # EXPORT EMISSION AND UNCERTAINTY PROPAGATION RESULTS TO EXCEL .XLSX (APPROACH 1)
    #===========================================================================
    
    ws = wb.create_sheet("results_u_pr_app1_" + tab_name)

    col_EM_BY =                   "C" #emission, base year
    col_EM_RY =                   "D" #emission, reporting year
    col_EM_u_lower =              "E" #emission uncertainty in percent (approx 2 sigmas), lower range compared to mean 
    col_EM_u_upper =              "F" #emission uncertainty in percent (approx 2 sigmas), upper range compared to mean 
    col_EM_contrib_lower =        "G" #contribution from this source category to inventory emission uncertainty, lower range compared to mean 
    col_EM_contrib_upper =        "H" #contribution from this source category to inventory emission uncertainty, upper range compared to mean
    col_sens_corr =               "I" #sensitivity if correlation is True (BY and RY are fully correlated)
    col_sens_not_corr =           "J" #sensitivity if correlation is False (BY and RY are not correlated)
    col_AD_contrib_trend_lower =  "K" #AD contribution to trend uncertainty, lower range compared to mean
    col_AD_contrib_trend_upper =  "L" #AD contribution to trend uncertainty, upper range compared to mean
    col_EF_contrib_trend_lower =  "M" #EF contribution to trend uncertainty, lower range compared to mean
    col_EF_contrib_trend_upper =  "N" #EF contribution to trend uncertainty, upper range compared to mean
    col_EM_contrib_trend_lower =  "O" #contribution from this source category to inventory trend uncertainty, lower range compared to mean 
    col_EM_contrib_trend_upper =  "P" #contribution from this source category to inventory trend uncertainty, upper range compared to mean
    col_end = col_EM_contrib_trend_upper #end of columns, to draw a line

    if routine == const.ROUTINE_NID:
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 24

    ws.column_dimensions[col_EM_BY].width = 10
    ws.column_dimensions[col_EM_RY].width = 10
    ws.column_dimensions[col_EM_u_lower].width = width_narrow_col
    ws.column_dimensions[col_EM_u_upper].width = width_narrow_col

    #Legend
    i_row = 1
    i_row_string = str(i_row)
        
    ws["A" + i_row_string] = \
    "Table {}: Uncertainty analysis of {} emissions, approach 1, for ".format(tab_name, comp_label) +\
    "{} and for the trend {}-{}{}. ".format(RY_string, BY_string, RY_string, text_LULUCF_and_indirect) +\
    "The uncertainties are given considering a 95% confidence interval " +\
    "and expressed as the distance from edge to mean, in percentage of the mean. " +\
    "AD: activity data; EF: emission factor; EM: emission; corr.: correlated. {}".format(dEM_text)

    ws.merge_cells("A" + i_row_string +":" + col_end + i_row_string)
    ws.row_dimensions[i_row].height = 60
    ws["A" + i_row_string].alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center", wrap_text=True)

    i_row += 1
    i_row_string = str(i_row)


    #write results, one line per routine
    count_row = int(-1)
    for i in index_mc_output:
        count_row += int(1)
        #write only processes with valid numeric entries
        #and where EM_BY or EM_RY is not zero.
        if df_pr["EM_is_num_BY"].iloc[i] or df_pr["EM_is_num_RY"].iloc[i]:
            if df_pr["reso_id"].iloc[i] != "" and (df_pr["reso_id"].iloc[i] != "Total" and df_pr["reso_id"].iloc[i] != "All resources"):
                code_reso_text = df_pr['proc_code'].iloc[i] + "; " + df_pr["reso_name"].iloc[i]
            else:
                code_reso_text = df_pr['proc_code'].iloc[i]
    
            if count_row == int(0) or (count_row > const.FORMAT_LINE_PER_PAGE_EXCEL and routine == const.ROUTINE_NID):
                count_row = 0
            
                #*********************************HEADER**********************************
                header_row_start = i_row

                #write letters A, B, C etc. according to IPCC formating Table 3.2
                ws["A" + i_row_string].value = "A" 
                ws["B" + i_row_string].value = "B" 
                ws[col_EM_BY + i_row_string].value = "C" 
                ws[col_EM_RY + i_row_string].value = "D" 
            
                ws[col_EM_u_lower + i_row_string].value = "G"
                ws.merge_cells(col_EM_u_lower + i_row_string +":" + col_EM_u_upper + i_row_string)
            
                ws[col_EM_contrib_lower + i_row_string] = "H"
                ws.merge_cells(col_EM_contrib_lower + i_row_string +":" + col_EM_contrib_upper + i_row_string)
            
                ws[col_sens_corr + i_row_string] = "I"
                ws[col_sens_not_corr + i_row_string] = "J"
            
                ws[col_AD_contrib_trend_lower + i_row_string] = "K"
                ws.merge_cells(col_AD_contrib_trend_lower + i_row_string +":" + col_AD_contrib_trend_upper + i_row_string)
            
                ws[col_EF_contrib_trend_lower + i_row_string] = "L"
                ws.merge_cells(col_EF_contrib_trend_lower + i_row_string +":" + col_EF_contrib_trend_upper + i_row_string)
            
                ws[col_EM_contrib_trend_lower + i_row_string] = "M" #total
                ws.merge_cells(col_EM_contrib_trend_lower + i_row_string +":" + col_EM_contrib_trend_upper + i_row_string)
                
                i_row += 1
                i_row_string = str(i_row)
                
                ws.row_dimensions[i_row].height = 60
                
                ws["A" + i_row_string] = nomenc_text
                ws.merge_cells("A" + i_row_string +":" + "A" + str(i_row+1))
                ws["B" + i_row_string] = comp_text
                ws.merge_cells("B" + i_row_string +":" + "B" + str(i_row+1))
                ws[col_EM_BY + i_row_string]= "Emissions " + BY_string
                ws[col_EM_RY + i_row_string]= "Emissions " + RY_string
            
                ws[col_EM_u_lower + i_row_string] = "Emission combined uncertainty {}".format(RY_string)
                ws.merge_cells(col_EM_u_lower + i_row_string +":" + col_EM_u_upper + i_row_string)
            
                ws[col_EM_contrib_lower + i_row_string] = "Category contribution to inventory variance {}".format(RY_string)
                ws.merge_cells(col_EM_contrib_lower + i_row_string +":" + col_EM_contrib_upper + i_row_string)
            
                ws[col_sens_corr + i_row_string] = "Sensitivi- ty if corr. (type A)"
                ws[col_sens_not_corr + i_row_string] = "Sensitivi- ty if not corr. (type B)"
            
                ws[col_AD_contrib_trend_lower + i_row_string] = "Contribution to inventory trend uncertainty from AD"
                ws.merge_cells(col_AD_contrib_trend_lower + i_row_string +":" + col_AD_contrib_trend_upper + i_row_string)
            
                ws[col_EF_contrib_trend_lower + i_row_string] = "Contribution to inventory trend uncertainty from EF"
                ws.merge_cells(col_EF_contrib_trend_lower + i_row_string +":" + col_EF_contrib_trend_upper + i_row_string)
            
                ws[col_EM_contrib_trend_lower + i_row_string] = "Contribution to inventory trend uncertainty from EM" #total
                ws.merge_cells(col_EM_contrib_trend_lower + i_row_string +":" + col_EM_contrib_trend_upper + i_row_string)
                        
                i_row += 1
                i_row_string = str(i_row)
            
                ws[col_EM_u_lower + i_row_string] = "(-)%"
                ws[col_EM_u_upper + i_row_string] = "(+)%"
                ws[col_EM_contrib_lower + i_row_string] = "(-)%"
                ws[col_EM_contrib_upper + i_row_string] = "(+)%"
                ws[col_AD_contrib_trend_lower + i_row_string] = "(-)%"
                ws[col_AD_contrib_trend_upper + i_row_string] = "(+)%"
                ws[col_EF_contrib_trend_lower + i_row_string] = "(-)%"
                ws[col_EF_contrib_trend_upper + i_row_string] = "(+)%"
                ws[col_EM_contrib_trend_lower + i_row_string] = "(-)%"
                ws[col_EM_contrib_trend_upper + i_row_string] = "(+)%"
            
                for row in ws.iter_rows(min_col = excel_columns.index("A") + 1, 
                                        min_row= header_row_start, 
                                        max_col= excel_columns.index(col_end) + 1, 
                                        max_row= i_row):
                    for cell in row:
                        cell.border = medium_border
                        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
            
                i_row += 1
                i_row_string = str(i_row)
            
                #******************************END OF HEADER******************************                 

            ws.row_dimensions[i_row].height = 13
            ws["A" + i_row_string].number_format = openpyxl.styles.numbers.FORMAT_TEXT
            ws["B" + i_row_string].number_format = openpyxl.styles.numbers.FORMAT_TEXT

            ws["A" + i_row_string] = code_reso_text                
            ws["B" + i_row_string] = df_pr["comp_name"].iloc[i]

            if df_pr['EM_status_BY'].iloc[i] == "ES":
                ws[col_EM_BY + i_row_string] = df_pr['EM_BY'].iloc[i]
                ws[col_EM_BY + i_row_string].number_format = const.FORMAT_VAL_EM #'##0.00'
            else:
                apply_style_non_numeric(df_pr['EM_status_BY'].iloc[i], ws[col_EM_BY + i_row_string])
    
            if df_pr['EM_status_RY'].iloc[i] == "ES":
                #emission for RY
                ws[col_EM_RY + i_row_string] = df_pr["EM_RY"].iloc[i]
                ws[col_EM_RY + i_row_string].number_format = const.FORMAT_VAL_EM #'##0.00''##0.00'

                #emission uncertainty for RY, in percent of the mean, 95% conf. interv.
                ws[col_EM_u_lower + i_row_string] = df_pr["EM_RY_pr_U_lower_p"].iloc[i]
                ws[col_EM_u_upper + i_row_string] = df_pr["EM_RY_pr_U_upper_p"].iloc[i]
                ws[col_EM_u_lower + i_row_string].number_format = const.FORMAT_VAL_U
                ws[col_EM_u_upper + i_row_string].number_format = const.FORMAT_VAL_U

                #emission contribution to inventory variance
                ws[col_EM_contrib_lower + i_row_string] = df_pr["EM_RY_pr_contrib_var_normed_lower"].iloc[i]
                ws[col_EM_contrib_upper + i_row_string] = df_pr["EM_RY_pr_contrib_var_normed_upper"].iloc[i]
                ws[col_EM_contrib_lower + i_row_string].number_format = const.FORMAT_VAL_3D
                ws[col_EM_contrib_upper + i_row_string].number_format = const.FORMAT_VAL_3D
                
            else:
                apply_style_non_numeric(df_pr['EM_status_RY'].iloc[i], ws[col_EM_RY + i_row_string])
                apply_style_non_numeric(df_pr['EM_status_RY'].iloc[i], ws[col_EM_u_lower + i_row_string])
                apply_style_non_numeric(df_pr['EM_status_RY'].iloc[i], ws[col_EM_u_upper + i_row_string])
                apply_style_non_numeric(df_pr['EM_status_RY'].iloc[i], ws[col_EM_contrib_lower + i_row_string])
                apply_style_non_numeric(df_pr['EM_status_RY'].iloc[i], ws[col_EM_contrib_upper + i_row_string])

            #EM contribution to trend variance, columns are always filled in             
            ws[col_EM_contrib_trend_lower + i_row_string] = df_pr["EM_trend_normed_pr_contrib_var_lower"].iloc[i]
            ws[col_EM_contrib_trend_upper + i_row_string] = df_pr["EM_trend_normed_pr_contrib_var_upper"].iloc[i]            
            ws[col_EM_contrib_trend_lower + i_row_string].number_format = const.FORMAT_VAL_CONTRIB
            ws[col_EM_contrib_trend_upper + i_row_string].number_format = const.FORMAT_VAL_CONTRIB
            
            
    
    
            #AD and EF uncertainty
            if not df_pr["import"].iloc[i]:
                apply_style_empty(ws[col_AD_contrib_trend_lower + i_row_string])
                apply_style_empty(ws[col_AD_contrib_trend_upper + i_row_string])
                apply_style_empty(ws[col_EF_contrib_trend_lower + i_row_string])
                apply_style_empty(ws[col_EF_contrib_trend_upper + i_row_string])
                apply_style_empty(ws[col_sens_corr + i_row_string])
                apply_style_empty(ws[col_sens_not_corr + i_row_string])
                
            else:
                #else, this is an import index
                i_in = df_in.index[(df_in["proc_id"] == df_pr["proc_id"].iloc[i]) 
                            & (df_in["reso_id"] == df_pr["reso_id"].iloc[i]) 
                            & (df_in["comp_id"] == df_pr["comp_id"].iloc[i])].tolist()
                                
                if len(i_in) == 0:
                    print("***********************")
                    print("match not found in input!")
                    print(df_pr["proc_id"].iloc[i], df_pr["reso_id"].iloc[i], df_pr["comp_id"].iloc[i])
                elif len(i_in) > 1:
                    print("too many matches found in input!")
                    print("***********************")
                    print(df_in["proc_id"].iloc[i_in], df_in["reso_id"].iloc[i_in], df_in["comp_id"].iloc[i_in], df_in["uEM_is_num_RY"].iloc[i_in], df_pr_AD_EF["AD_RY_pr_U_lower_p"].iloc[i_in])
                    print(df_pr["proc_id"].iloc[i], df_pr["reso_id"].iloc[i], df_pr["comp_id"].iloc[i])
                    
                i_in = i_in[0]

                #sensitivity, columns are always filled in
                ws[col_sens_corr + i_row_string] = df_pr_AD_EF["sens_corr"].iloc[i_in]
                ws[col_sens_not_corr + i_row_string] = df_pr_AD_EF["sens_no_corr"].iloc[i_in]
                ws[col_sens_corr + i_row_string].number_format = const.FORMAT_VAL_CONTRIB
                ws[col_sens_not_corr + i_row_string].number_format = const.FORMAT_VAL_CONTRIB
                    
                if df_in["uEM_is_num_RY"].iloc[i_in]:
                #input emission is not zero and this is a direct emission
                    apply_style_dEM(ws[col_AD_contrib_trend_lower + i_row_string])
                    apply_style_dEM(ws[col_AD_contrib_trend_upper + i_row_string])
                    apply_style_dEM(ws[col_EF_contrib_trend_lower + i_row_string])
                    apply_style_dEM(ws[col_EF_contrib_trend_upper + i_row_string])
                                
                else:
                    #print("writting values for AD and EF")
                    #input emission is not zero and input are given for AD and EF   
                    #AD contribution to trend variance, as computed from uncertainty propagation
                    ws[col_AD_contrib_trend_lower + i_row_string] = df_pr_AD_EF["AD_trend_normed_pr_contrib_var_lower"].iloc[i_in]
                    ws[col_AD_contrib_trend_upper + i_row_string] = df_pr_AD_EF["AD_trend_normed_pr_contrib_var_upper"].iloc[i_in]
    
                    #EF contribution to trend variance, as computed from uncertainty propagation
                    ws[col_EF_contrib_trend_lower + i_row_string] = df_pr_AD_EF["EF_trend_normed_pr_contrib_var_lower"].iloc[i_in]
                    ws[col_EF_contrib_trend_upper + i_row_string] = df_pr_AD_EF["EF_trend_normed_pr_contrib_var_upper"].iloc[i_in]
                    
                    ws[col_AD_contrib_trend_lower + i_row_string].number_format = const.FORMAT_VAL_CONTRIB
                    ws[col_AD_contrib_trend_upper + i_row_string].number_format = const.FORMAT_VAL_CONTRIB
                    ws[col_EF_contrib_trend_lower + i_row_string].number_format = const.FORMAT_VAL_CONTRIB
                    ws[col_EF_contrib_trend_upper + i_row_string].number_format = const.FORMAT_VAL_CONTRIB

    
            for row in ws.iter_rows(min_col = 1, 
                                    min_row = i_row, 
                                    max_col= excel_columns.index(col_end) + 1, 
                                    max_row= i_row):
                for cell in row:
                    cell.border = thin_border
            
            i_row += 1
            i_row_string = str(i_row)

                            
    #***************WRITE INVENTORY SUM******************************

    #**************inventory variance********************************
    ws["A" + i_row_string] = "Total"
    ws.merge_cells(col_EM_u_lower + i_row_string +":" + col_EM_u_upper + i_row_string)
    ws[col_EM_contrib_lower + i_row_string] = df_pr["EM_RY_pr_contrib_var_normed_lower"].iloc[index_pr_total] 
    ws[col_EM_contrib_upper + i_row_string] = df_pr["EM_RY_pr_contrib_var_normed_upper"].iloc[index_pr_total] 

    ws.merge_cells(col_AD_contrib_trend_lower + i_row_string +":" + col_EF_contrib_trend_upper + i_row_string)

    ws[col_EM_contrib_trend_lower + i_row_string] = df_pr["EM_trend_normed_pr_contrib_var_lower"].iloc[index_pr_total] 
    ws[col_EM_contrib_trend_upper + i_row_string] = df_pr["EM_trend_normed_pr_contrib_var_upper"].iloc[index_pr_total] 
    
    for row in ws.iter_rows(min_col = excel_columns.index("A") + 1, 
                            min_row= i_row, 
                            max_col= excel_columns.index(col_end) + 1, 
                            max_row= i_row):
        for cell in row:
            cell.border = thin_border
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="8DB4E2")
            cell.number_format = const.FORMAT_VAL_TOTAL

    #**************inventory uncertainty********************************    
    i_row += 1
    i_row_string = str(i_row)
    ws.row_dimensions[i_row].height = 30
    ws["A" + i_row_string] = "Total"
    ws[col_EM_BY + i_row_string] = df_pr['EM_BY'].iloc[index_pr_total] 
    ws[col_EM_RY + i_row_string] = df_pr['EM_RY'].iloc[index_pr_total]

    ws.merge_cells(col_EM_u_lower + i_row_string +":" + col_EM_u_upper + i_row_string)
    ws[col_EM_u_lower + i_row_string] = "Emissions {} uncertainty (%):".format(RY_string)
    ws[col_EM_u_lower + i_row_string].alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws[col_EM_contrib_lower + i_row_string] = df_pr["EM_RY_pr_U_lower_p"].iloc[index_pr_total]
    ws[col_EM_contrib_upper + i_row_string] = df_pr["EM_RY_pr_U_upper_p"].iloc[index_pr_total]

    ws.merge_cells(col_AD_contrib_trend_lower + i_row_string +":" + col_EF_contrib_trend_upper + i_row_string)
    ws[col_AD_contrib_trend_lower + i_row_string] = "Trend uncertainty (%):"
    ws[col_EM_contrib_trend_lower + i_row_string] = df_pr["EM_trend_normed_pr_U_lower_p"].iloc[index_pr_total] 
    ws[col_EM_contrib_trend_upper + i_row_string] = df_pr["EM_trend_normed_pr_U_upper_p"].iloc[index_pr_total] 

    for row in ws.iter_rows(min_col = excel_columns.index("A") + 1, 
                            min_row= i_row, 
                            max_col= excel_columns.index(col_end) + 1, 
                            max_row= i_row):
        for cell in row:
            cell.border = thin_border
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="8DB4E2")
            cell.number_format = const.FORMAT_VAL_TOTAL
    
            
    #***************END OF WRITE INVENTORY SUM******************************


    #===========================================================================
    # EXPORT EMISSION AND MONTE CARLO UNCERTAINTY RESULTS TO EXCEL .XLSX (APPROACH 2)
    #===========================================================================
    #write results in Excel table
    #Format: cf. IPCC guidelines 2006 Chap. 3, Table 3.3
    #"General reporting table for uncertainties"
    #this is not the same table as for Uncertainties Approach 1
    

        
    #ws = wb.active #active sheet
    ws = wb.create_sheet("results_u_mc_app2_" + tab_name)
        
    #set width of columns
    if routine == const.ROUTINE_NID:
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 6
    ws.column_dimensions["F"].width = 6
    ws.column_dimensions["G"].width = 6
    ws.column_dimensions["H"].width = 6
    ws.column_dimensions["I"].width = 6
    ws.column_dimensions["J"].width = 6
    
    col_EM_var_contrib = "K"
    col_trend_normed_mean = "L" #"O" #mean of trend normalised by sum of BY_EM, trend contribution to inventory trend
    col_trend_normed_u_lower = "M" #"P"
    col_trend_normed_u_upper = "N" #"Q"
    col_end = col_trend_normed_u_upper #col_trend_normed_var
    c_end = excel_columns.index(col_end)
    
    #extra columns to write extra values, not required for reporting
    col_EM_BY_mc_2stddev_p = "P"
    col_EM_RY_mc_2stddev_p = "Q"
    #col_EM_BY_RY_corr = "R"
    
    i_row = 1
    i_row_string = str(i_row)
    
    #Write legend of table    
    ws["A" + i_row_string]= \
    "Table {}: Uncertainty analysis of {} emissions, approach 2, for ".format(tab_name, comp_label) +\
    "{} and for the trend {}-{}{}. {} ".format(RY_string, BY_string, RY_string, text_LULUCF_and_indirect, dEM_text) +\
    "Monte Carlo simulations " +\
    "were run {} times. The reported uncertainties correspond to the borders ".format(no_mc) + \
    "of the narrowest {:.1f}% confidence interval. ".format(const.P_DIST*float(100)) + \
    "Contributions to inventory trend (mean, uncertainties, columns I and J) " +\
    "are values normalised by the total inventory base year emission."
    
    ws.merge_cells("A" + i_row_string + ":" + col_end +i_row_string)
    ws.row_dimensions[i_row].height = 80
    ws["A" + i_row_string].alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    i_row += 1
    i_row_string = str(i_row)
    #starting_row = True
    #split_idx = 0
                    
    #write results, one line per routine
    count_row = int(-1)
    for i in index_mc_output: #range(no_nomenc_in):
        count_row += int(1)
        #write only processes with valid numeric entries
        #and where EM_BY or EM_RY is not zero.
        if df_mc["EM_is_num_BY"].iloc[i] or df_mc["EM_is_num_RY"].iloc[i]:
            if df_mc["reso_id"].iloc[i] != "" and (df_mc["reso_id"].iloc[i] != "Total" and df_mc["reso_id"].iloc[i] != "All resources"):
                code_reso_text = df_mc['proc_code'].iloc[i] + "; " + df_mc["reso_name"].iloc[i]
            else:
                code_reso_text = df_mc['proc_code'].iloc[i]
    
            if count_row == int(0) or (count_row > const.FORMAT_LINE_PER_PAGE_EXCEL and routine == const.ROUTINE_NID):
                count_row = 0
            
                #*********************************HEADER**********************************
                header_row_start = i_row
                #add lettering of columns according to IPCC 2006, VOl 1, Table 3.3 (for Monte Carlo)
                ws["A" + i_row_string].value = "A"
                ws["B" + i_row_string].value = "B"
                ws["C" + i_row_string].value = "C"
                ws["D" + i_row_string].value = "D"
                
                ws["E" + i_row_string].value = "E"
                ws.merge_cells("E" + i_row_string + ":F" +i_row_string)
                
                ws["G" + i_row_string] = "F"
                ws.merge_cells("G" + i_row_string + ":H" +i_row_string)
            
                ws["I" + i_row_string] = "G"
                ws.merge_cells("I" + i_row_string + ":J" +i_row_string)
                
                ws[col_EM_var_contrib + i_row_string] = "H"
            
                #ws.merge_cells(col_trend_u_lower + i_row_string + ":" + col_trend_u_upper + i_row_string)
                ws[col_trend_normed_mean + i_row_string] = "I"
                ws[col_trend_normed_u_lower + i_row_string] = "J"
                ws.merge_cells(col_trend_normed_u_lower + i_row_string + ":" + col_trend_normed_u_upper + i_row_string)
                
                i_row += 1
                i_row_string = str(i_row)
                ws.row_dimensions[i_row].height = 72
            
                #assign data to cells
                ws["A" + i_row_string]= nomenc_text
                ws.merge_cells("A" + i_row_string + ":A" +str(i_row + 1) )
            
                ws["B" + i_row_string] = comp_text
                ws.merge_cells("B" + i_row_string + ":B" +str(i_row + 1) )
            
                ws["C" + i_row_string]= "Emissions " + BY_string
                ws["D" + i_row_string]= "Emissions " + RY_string
                ws["E" + i_row_string]= "Activity data uncertainty " + RY_string
                ws.merge_cells("E" + i_row_string + ":F" +i_row_string)
            
                ws["G" + i_row_string] = "Emission factor uncertainty " + RY_string
                ws.merge_cells("G" + i_row_string + ":H" +i_row_string)
            
                ws["I" + i_row_string] = "Emission combined uncertainty " + RY_string
                ws.merge_cells("I" + i_row_string + ":J" +i_row_string)
            
                ws[col_EM_var_contrib + i_row_string] = "Emission contri- bution to variance " + RY_string
                            
                ws[col_trend_normed_mean    + i_row_string] = "Contri- bution to trend"
                ws[col_trend_normed_u_lower + i_row_string] = "Contribution to uncertainty of trend"
                ws.merge_cells(col_trend_normed_u_lower + i_row_string + ":" + col_trend_normed_u_upper + i_row_string)
                
                ws[col_EM_BY_mc_2stddev_p + i_row_string] = "EM BY: 2 std dev %"
                ws[col_EM_RY_mc_2stddev_p + i_row_string] = "EM RY: 2 std dev %"
                #ws[col_EM_BY_RY_corr + i_row_string] = "EM BY RY corr"
            
                
                i_row += 1
                i_row_string = str(i_row)
                
                ws["C" + i_row_string] = unit_string
                ws["D" + i_row_string] = unit_string
                ws["E" + i_row_string] = "(-)%"
                ws["F" + i_row_string] = "(+)%"
                ws["G" + i_row_string] = "(-)%"
                ws["H" + i_row_string] = "(+)%"
                ws["I" + i_row_string] = "(-)%"
                ws["J" + i_row_string] = "(+)%"
                ws[col_EM_var_contrib + i_row_string] = "Fraction"
            
                ws[col_trend_normed_mean    + i_row_string] = "%"
                ws[col_trend_normed_u_lower + i_row_string] = "(-)%"
                ws[col_trend_normed_u_upper + i_row_string] = "(+)%"
                        
            
                for row in ws.iter_rows(min_col = 1, 
                                        min_row = header_row_start, 
                                        max_col= c_end + 1, 
                                        max_row= i_row):
                    for cell in row:
                        cell.border = medium_border
                        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
            
                i_row += 1
                i_row_string = str(i_row)
            
                #******************************END OF HEADER******************************            
            
 
            ws.row_dimensions[i_row].height = 13
            ws["A" + i_row_string].number_format = openpyxl.styles.numbers.FORMAT_TEXT
            ws["B" + i_row_string].number_format = openpyxl.styles.numbers.FORMAT_TEXT
    
            ws["A" + i_row_string] = code_reso_text            
            ws["B" + i_row_string] = df_mc["comp_name"].iloc[i]
    
            if df_mc['EM_status_BY'].iloc[i] == "ES":
                ws["C" + i_row_string] = df_mc['EM_BY'].iloc[i]
                ws["C" + i_row_string].number_format = const.FORMAT_VAL_EM #'##0.00'
            else:
                apply_style_non_numeric(df_mc['EM_status_BY'].iloc[i], ws["C" + i_row_string])
    
            if df_mc['EM_status_RY'].iloc[i] == "ES":
                ws["D" + i_row_string] = df_mc['EM_RY'].iloc[i]
                ws["D" + i_row_string].number_format = const.FORMAT_VAL_EM #'##0.00''##0.00'
            else:
                apply_style_non_numeric(df_mc['EM_status_RY'].iloc[i], ws["D" + i_row_string])
                apply_style_non_numeric(df_mc['EM_status_RY'].iloc[i], ws["E" + i_row_string])
                apply_style_non_numeric(df_mc['EM_status_RY'].iloc[i], ws["F" + i_row_string])
                apply_style_non_numeric(df_mc['EM_status_RY'].iloc[i], ws["G" + i_row_string])
                apply_style_non_numeric(df_mc['EM_status_RY'].iloc[i], ws["H" + i_row_string])
                apply_style_non_numeric(df_mc['EM_status_RY'].iloc[i], ws["I" + i_row_string])
                apply_style_non_numeric(df_mc['EM_status_RY'].iloc[i], ws["J" + i_row_string])
                #apply_style_non_numeric(df_mc['EM_status_RY'].iloc[i], ws["K" + i_row_string])
    
            for row in ws.iter_rows(min_col = excel_columns.index("E") + 1, 
                                    min_row = i_row, 
                                    max_col= excel_columns.index("J") + 1, 
                                    max_row= i_row):
                for cell in row:
                    cell.number_format = const.FORMAT_VAL_U #'##0'
    
    
            #AD and EF uncertainty
            if not df_mc["import"].iloc[i]:
                apply_style_empty(ws["E" + i_row_string])
                apply_style_empty(ws["F" + i_row_string])
                apply_style_empty(ws["G" + i_row_string])
                apply_style_empty(ws["H" + i_row_string])
                
            else:
                #else, this is an import index
                i_in = df_in.index[(df_in["proc_id"] == df_mc["proc_id"].iloc[i]) 
                            & (df_in["reso_id"] == df_mc["reso_id"].iloc[i]) 
                            & (df_in["comp_id"] == df_mc["comp_id"].iloc[i])].tolist()
                
                
                if len(i_in) == 0:
                    print("***********************")
                    print("match not found in input!")
                    print(df_mc["proc_id"].iloc[i], df_mc["reso_id"].iloc[i], df_mc["comp_id"].iloc[i])
                elif len(i_in) > 1:
                    print("too many matches found in input!")
                    print("***********************")
                    print(df_in["proc_id"].iloc[i_in], df_in["reso_id"].iloc[i_in], df_in["comp_id"].iloc[i_in], df_in["uEM_is_num_RY"].iloc[i_in], df_mc_AD_EF["AD_RY_mc_U_lower_p"].iloc[i_in])
                    print(df_mc["proc_id"].iloc[i], df_mc["reso_id"].iloc[i], df_mc["comp_id"].iloc[i])
                    
                i_in = i_in[0]
                
                                
                if df_mc["EM_status_RY"].iloc[i] != "ES":
                    #input emission is zero, write notation key instead
                    apply_style_non_numeric(df_mc["EM_status_RY"].iloc[i], ws["E" + i_row_string])
                    apply_style_non_numeric(df_mc["EM_status_RY"].iloc[i], ws["F" + i_row_string])
                    apply_style_non_numeric(df_mc["EM_status_RY"].iloc[i], ws["G" + i_row_string])
                    apply_style_non_numeric(df_mc["EM_status_RY"].iloc[i], ws["H" + i_row_string])
    
                
                elif df_in["uEM_is_num_RY"].iloc[i_in]:
                #input emission is not zero and this is a direct emission
                    apply_style_dEM(ws["E" + i_row_string])
                    apply_style_dEM(ws["F" + i_row_string])
                    apply_style_dEM(ws["G" + i_row_string])
                    apply_style_dEM(ws["H" + i_row_string])
        
                        
                else:
                    #print("writting values for AD and EF")
                    #input emission is not zero and input are given for AD and EF   
                    #AD uncertainty as generated from mc simulations
                    ws["E" + i_row_string] = df_mc_AD_EF["AD_RY_mc_U_lower_p"].iloc[i_in]
                    ws["F" + i_row_string] = df_mc_AD_EF["AD_RY_mc_U_upper_p"].iloc[i_in]
    
                    #EF uncertainty as generated from mc simulations
                    ws["G" + i_row_string] = df_mc_AD_EF["EF_RY_mc_U_lower_p"].iloc[i_in]
                    ws["H" + i_row_string] = df_mc_AD_EF["EF_RY_mc_U_upper_p"].iloc[i_in]
    
            #EM uncertainty from mc simulations and EM contribution to variance
            if df_mc['EM_status_RY'].iloc[i] != "ES":
                apply_style_non_numeric(df_mc['EM_status_RY'].iloc[i], ws["I" + i_row_string])
                apply_style_non_numeric(df_mc['EM_status_RY'].iloc[i], ws["J" + i_row_string])
                ws[col_EM_var_contrib + i_row_string] = float(0)

            elif not pd.isnull(df_mc["EM_RY_mc_mean"].iloc[i]):
                #if the simulated emission is zero because the input is zero, write zero.
                ws["I" + i_row_string] = df_mc["EM_RY_mc_U_lower_p"].iloc[i]
                ws["J" + i_row_string] = df_mc["EM_RY_mc_U_upper_p"].iloc[i]
                ws[col_EM_var_contrib + i_row_string] = df_mc["EM_RY_mc_var_normed"].iloc[i]
                
                ws[col_EM_BY_mc_2stddev_p + i_row_string] = df_mc["EM_BY_mc_2stddev_p"].iloc[i]
                ws[col_EM_RY_mc_2stddev_p + i_row_string] = df_mc["EM_RY_mc_2stddev_p"].iloc[i]
                #ws[col_EM_BY_RY_corr + i_row_string] = df_mc["EM_BY_RY_mc_sensitivity"].iloc[i]
                                
            else:
                ws["I" + i_row_string] = float(0)
                ws["J" + i_row_string] = float(0)
                ws[col_EM_var_contrib + i_row_string] = float(0)
        
            ws[col_EM_var_contrib + i_row_string].number_format = const.FORMAT_VAL_CONTRIB
     
                    
            #fill in uncertainty for trend expressed in percent of total trend uncertainty!!!
            #so this is not the uncertainty asociated with each trend
            #the sum of this input should be the value of uncertainty for the overall trend
            #(trend of all EM summed up)
            ws[col_trend_normed_mean    + i_row_string] = df_mc["EM_trend_normed_mc_mean"].iloc[i] 
            ws[col_trend_normed_mean    + i_row_string].number_format = const.FORMAT_VAL_CONTRIB #'##0.000'
                        
            ws[col_trend_normed_u_lower + i_row_string] = df_mc["EM_trend_normed_mc_U_lower_p"].iloc[i] 
            ws[col_trend_normed_u_lower + i_row_string].number_format = const.FORMAT_VAL_CONTRIB #'##0.000'
            ws[col_trend_normed_u_upper + i_row_string] = df_mc["EM_trend_normed_mc_U_upper_p"].iloc[i] 
            ws[col_trend_normed_u_upper + i_row_string].number_format = const.FORMAT_VAL_CONTRIB #'##0.000'
    
    
            for row in ws.iter_rows(min_col = 1, 
                                    min_row = i_row, 
                                    max_col= c_end + 1, 
                                    max_row= i_row):
                for cell in row:
                    cell.border = thin_border
            
            i_row += 1
            i_row_string = str(i_row)
        
    #write average results, this is the last line
    

    ws["A"+ i_row_string] = "Total, Monte Carlo simulations"
    ws["A"+ i_row_string].alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[i_row].height = 30
    ws.merge_cells("A" + i_row_string + ":" + "B" + i_row_string)
    ws["B"+ i_row_string].border = thin_border
    ws["C"+ i_row_string] = df_mc["EM_BY_mc_mean"].iloc[index_mc_total] 
    ws["D"+ i_row_string] = df_mc["EM_RY_mc_mean"].iloc[index_mc_total] 
    ws["I"+ i_row_string] = df_mc["EM_RY_mc_U_lower_p"].iloc[index_mc_total] 
    ws["J"+ i_row_string] = df_mc["EM_RY_mc_U_upper_p"].iloc[index_mc_total] 
    ws[col_EM_var_contrib    + i_row_string] = np.nansum(df_mc["EM_BY_mc_var_normed"].iloc[index_mc_output]) #this must be 1
    #The following is not 1, this is only the variance of the inventory sum:
    #df_mc["EM_RY_mc_var_normed"].iloc[index_mc_total]
    
    ws[col_trend_normed_mean    + i_row_string] = df_mc["EM_trend_normed_mc_mean"].iloc[index_mc_total]
    ws[col_trend_normed_u_lower + i_row_string] = df_mc["EM_trend_normed_mc_U_lower_p"].iloc[index_mc_total] 
    ws[col_trend_normed_u_upper + i_row_string] = df_mc["EM_trend_normed_mc_U_upper_p"].iloc[index_mc_total] 
    
    for row in ws.iter_rows(min_col = 1, 
                            min_row = i_row, 
                            max_col= c_end + 1, 
                            max_row= i_row):
        for cell in row:
            cell.border = thin_border
            cell.number_format = const.FORMAT_VAL_TOTAL #'##0.0'
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="8DB4E2")
            cell.alignment = openpyxl.styles.Alignment(vertical="center", wrap_text=True)
    
    
    i_row += 1
    i_row_string = str(i_row)
    
    
    #Total emissions over all processes, compounds, resources, according to the official inventory
    ws["A"+ i_row_string] = "Total, inventory"
    ws["A"+ i_row_string].alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells("A" + i_row_string + ":" + "B" + i_row_string)
    ws["B"+ i_row_string].border = thin_border
    ws["C"+ i_row_string] = df_mc["EM_BY"].iloc[index_mc_total]
    ws["D"+ i_row_string] = df_mc["EM_RY"].iloc[index_mc_total]
    ws[col_trend_normed_mean    + i_row_string] = (df_mc["EM_RY"].iloc[index_mc_total] - df_mc["EM_BY"].iloc[index_mc_total])/df_mc["EM_BY"].iloc[index_mc_total] * float(100.0) #trend_inventory
    
    for row in ws.iter_rows(min_col = 1, 
                            min_row = i_row, 
                            max_col= c_end + 1, 
                            max_row= i_row):
        for cell in row:
            cell.border = thin_border
            cell.number_format = const.FORMAT_VAL_TOTAL #'##0.0'
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="8DB4E2")



    #=================================================================
    #WRITE RESULTS FOR EACH SECTOR AND EACH COMPOUND
    #=================================================================
    
    for i_tab in range(2):
        if i_tab == 0:            
            #=================================================================
            #WRITE RESULTS FOR EACH SECTOR
            #=================================================================
            ws = wb.create_sheet("results_pr_mc_per_sector")            
            category_text = "Source category"
            col_cat_name = "proc_code_name"
            index_mc_list = index_mc_proc_sector_total
            index_mc_list_with_total = index_mc_proc_sector_total + [index_mc_total]
            index_pr_list = index_pr_proc_sector_total
            index_pr_list_with_total = index_pr_proc_sector_total + [index_pr_total]
            #print("index_mc_list sector ")
            #print(index_mc_list)

        elif i_tab == 1 and routine == const.ROUTINE_NID:
            #=================================================================
            #WRITE RESULTS FOR EACH COMPOUND IF APPLICABLE
            #=================================================================
            ws = wb.create_sheet("results_pr_mc_per_comp")                   
            category_text = "Gas"
            col_cat_name = "comp_name"
            index_mc_list = index_mc_comp_total
            index_mc_list_with_total = index_mc_comp_total + [index_mc_total]
            index_pr_list = index_pr_comp_total
            index_pr_list_with_total = index_pr_comp_total + [index_pr_total]
            #print("index_mc_list compound ")
            #print(index_mc_list)


        if i_tab == 0 or (i_tab == 1 and routine == const.ROUTINE_NID):
        
            col_cat = "A" #category: process or compound
            col_EM_BY = "B"
            col_EM_BY_U_lower_p = "C" #in percent of the mean
            col_EM_BY_U_upper_p = "D"
            col_EM_BY_U_contrib_p = "E"
            #col_em_u_mean = "E"
            col_EM_BY_end = col_EM_BY_U_contrib_p
            col_EM_RY = "F"
            col_EM_RY_U_lower_p = "G" 
            col_EM_RY_U_upper_p = "H"
            col_EM_RY_U_contrib_p = "I"
            col_EM_RY_end = col_EM_RY_U_contrib_p
            col_trend = "J"
            col_trend_U_lower_p = "K"
            col_trend_U_upper_p = "L"
            col_trend_U_contrib_p = "M"
            col_trend_end = col_trend_U_contrib_p
            col_end_str = col_trend_end
            
            col_end = [ii for ii in range(len(excel_columns)) if excel_columns[ii] == col_trend_end]
            col_end = col_end[0]
            
            ws.column_dimensions[col_cat].width = 30
            
            
            #******************************HEADER******************************   
            i_row = 1
            i_row_string = str(i_row)
            header_row_start = i_row
            #ws.row_dimensions[i_row].height = 45
        
            #assign data to cells
            ws[col_cat + i_row_string] = category_text
            ws.merge_cells(col_cat + i_row_string + ":" + col_cat + str(i_row + 1) )
        
            ws[col_EM_BY + i_row_string]= "Emissions {}".format(BY_string)
            ws.merge_cells(col_EM_BY + i_row_string + ":" + col_EM_BY_end + i_row_string)
        
            ws[col_EM_RY + i_row_string]= "Emissions {}".format(RY_string)
            ws.merge_cells(col_EM_RY + i_row_string + ":" + col_EM_RY_end + i_row_string)
        
            ws[col_trend + i_row_string]= "Contribution to trend {}-{}".format(BY_string, RY_string)
            ws.merge_cells(col_trend + i_row_string + ":" + col_trend_end + i_row_string)
            
            i_row += 1
            i_row_string = str(i_row)
            
            ws[col_EM_BY + i_row_string] = "Value " + unit_string
            ws[col_EM_BY_U_lower_p + i_row_string] = "U(-)%"
            ws[col_EM_BY_U_upper_p + i_row_string] = "U(+)%"
            ws[col_EM_BY_U_contrib_p + i_row_string] = "Contrib. fraction"
        
            ws[col_EM_RY + i_row_string] = "Value " + unit_string
            ws[col_EM_RY_U_lower_p + i_row_string] = "U(-)%"
            ws[col_EM_RY_U_upper_p + i_row_string] = "U(+)%"
            ws[col_EM_RY_U_contrib_p + i_row_string] = "Contrib. fraction"
        
            ws[col_trend + i_row_string] = "Value %"
            ws[col_trend_U_lower_p + i_row_string] = "U(-)%"
            ws[col_trend_U_upper_p + i_row_string] = "U(+)%"
            ws[col_trend_U_contrib_p + i_row_string] = "Contrib. fraction"
        
        
            for row in ws.iter_rows(min_col = 1, 
                                    min_row = header_row_start, 
                                    max_col= col_end + 1, 
                                    max_row= i_row):
                for cell in row:
                    cell.border = medium_border
                    cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
        
            i_row += 1
            i_row_string = str(i_row)     
            #******************************END OF HEADER***************************
    
            #******************************HEADER FOR PR***************************
            ws[col_cat + i_row_string] = "Uncertainty propagation (approach 1)"
            ws.merge_cells(col_cat + i_row_string + ":" + col_end_str + i_row_string)
    
            for row in ws.iter_rows(min_col = 1, 
                                    min_row = i_row, 
                                    max_col= col_end + 1, 
                                    max_row= i_row):
                for cell in row:
                    cell.border = medium_border
                    cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            i_row += 1
            i_row_string = str(i_row)
            #******************************END OF HEADER FOR PR********************
    
            for i in index_pr_list_with_total:
                #category
                if i == index_pr_total:
                    ws[col_cat + i_row_string] = "Total, inventory"
                    ws[col_cat + i_row_string].alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center", wrap_text=True)               
                else:        
                    ws[col_cat + i_row_string] = df_pr[col_cat_name].iloc[i]
                
                if df_pr['EM_status_BY'].iloc[i] == "ES":
                    #BY emission as per inventory
                    apply_number_format(df_pr["EM_BY"].iloc[i], ws[col_EM_BY + i_row_string], int(2))
                    #BY U (-) mc
                    apply_number_format(df_pr["EM_BY_pr_U_lower_p"].iloc[i], ws[col_EM_BY_U_lower_p + i_row_string], int(2))
                    #BY U (+) mc 
                    apply_number_format(df_pr["EM_BY_pr_U_upper_p"].iloc[i], ws[col_EM_BY_U_upper_p + i_row_string], int(2))
                    #BY contribution to total U mc
                    if i == index_pr_total:
                        value = np.nansum(df_pr["EM_BY_pr_var_normed"].iloc[index_pr_list])
                    else:
                        value = df_pr["EM_BY_pr_var_normed"].iloc[i]
                    apply_number_format(value, ws[col_EM_BY_U_contrib_p + i_row_string], int(2))
                else:
                    apply_style_non_numeric(df_pr['EM_status_BY'].iloc[i], ws[col_EM_BY + i_row_string])
                    apply_style_non_numeric(df_pr['EM_status_BY'].iloc[i], ws[col_EM_BY_U_lower_p + i_row_string])
                    apply_style_non_numeric(df_pr['EM_status_BY'].iloc[i], ws[col_EM_BY_U_upper_p + i_row_string])
                    apply_style_non_numeric(df_pr['EM_status_BY'].iloc[i], ws[col_EM_BY_U_contrib_p + i_row_string])
    
                if df_pr['EM_status_RY'].iloc[i] == "ES":
                    #RY emission as per inventory
                    apply_number_format(df_pr["EM_RY"].iloc[i], ws[col_EM_RY + i_row_string], int(2))
                    #RY U (-) mc
                    apply_number_format(df_pr["EM_RY_pr_U_lower_p"].iloc[i], ws[col_EM_RY_U_lower_p + i_row_string], int(2))
                    #RY U (+) mc
                    apply_number_format(df_pr["EM_RY_pr_U_upper_p"].iloc[i], ws[col_EM_RY_U_upper_p + i_row_string], int(2))
                    #RY contribution to total U mc
                    if i == index_pr_total:
                        value = np.nansum(df_pr["EM_RY_pr_var_normed"].iloc[index_pr_list])
                    else:
                        value = df_pr["EM_RY_pr_var_normed"].iloc[i]
                    apply_number_format(value, ws[col_EM_RY_U_contrib_p + i_row_string], int(2))
                else:
                    apply_style_non_numeric(df_pr['EM_status_RY'].iloc[i], ws[col_EM_RY + i_row_string])
                    apply_style_non_numeric(df_pr['EM_status_RY'].iloc[i], ws[col_EM_RY_U_lower_p + i_row_string])
                    apply_style_non_numeric(df_pr['EM_status_RY'].iloc[i], ws[col_EM_RY_U_upper_p + i_row_string])
                    apply_style_non_numeric(df_pr['EM_status_RY'].iloc[i], ws[col_EM_RY_U_contrib_p + i_row_string])
    
                if df_pr['EM_status_BY'].iloc[i] == "ES" or df_pr['EM_status_RY'].iloc[i] == "ES":
                    apply_number_format(df_pr["EM_trend_normed"].iloc[i], ws[col_trend + i_row_string], int(3))
                    apply_number_format(df_pr["EM_trend_normed_pr_U_lower_p"].iloc[i], ws[col_trend_U_lower_p + i_row_string], int(2))
                    apply_number_format(df_pr["EM_trend_normed_pr_U_upper_p"].iloc[i], ws[col_trend_U_upper_p + i_row_string], int(2))
                    if i == index_pr_total:
                        value = np.nansum(df_pr["EM_trend_normed_pr_var_normed"].iloc[index_pr_list])
                    else:
                        value = df_pr["EM_trend_normed_pr_var_normed"].iloc[i]
                    apply_number_format(value, ws[col_trend_U_contrib_p + i_row_string], int(2))
                else: #both are non-numeric
                    apply_style_non_numeric(df_pr['EM_status_RY'].iloc[i], ws[col_trend + i_row_string])
                    apply_style_non_numeric(df_pr['EM_status_RY'].iloc[i], ws[col_trend_U_lower_p + i_row_string])
                    apply_style_non_numeric(df_pr['EM_status_RY'].iloc[i], ws[col_trend_U_upper_p + i_row_string])
                    apply_style_non_numeric(df_pr['EM_status_RY'].iloc[i], ws[col_trend_U_contrib_p + i_row_string])
        
                for row in ws.iter_rows(min_col = 1, 
                                        min_row = i_row, 
                                        max_col= col_end + 1, #excel_columns
                                        max_row= i_row):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = openpyxl.styles.Alignment(vertical="center", wrap_text=True)
                        if i == index_pr_total:
                            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="8DB4E2")
        
                i_row += 1
                i_row_string = str(i_row) 
    
            
            #******************************HEADER FOR MC***************************
            ws[col_cat + i_row_string] = "Monte Carlo simulations (approach 2)"
            ws.merge_cells(col_cat + i_row_string + ":" + col_end_str + i_row_string)
    
            for row in ws.iter_rows(min_col = 1, 
                                    min_row = i_row, 
                                    max_col= col_end + 1, 
                                    max_row= i_row):
                for cell in row:
                    cell.border = medium_border
                    cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            i_row += 1
            i_row_string = str(i_row)
            #******************************END OF HEADER FOR MC********************
    
            
            for i in index_mc_list_with_total:
                #category
                if i == index_mc_total:
                    ws[col_cat + i_row_string] = "Total, inventory"
                    ws[col_cat + i_row_string].alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center", wrap_text=True)               
                else:        
                    ws[col_cat + i_row_string] = df_mc[col_cat_name].iloc[i]
                
                if df_mc['EM_status_BY'].iloc[i] == "ES":
                    #BY emission as per inventory
                    apply_number_format(df_mc["EM_BY"].iloc[i], ws[col_EM_BY + i_row_string], int(2))
                    #BY U (-) mc
                    apply_number_format(df_mc["EM_BY_mc_U_lower_p"].iloc[i], ws[col_EM_BY_U_lower_p + i_row_string], int(2))
                    #BY U (+) mc 
                    apply_number_format(df_mc["EM_BY_mc_U_upper_p"].iloc[i], ws[col_EM_BY_U_upper_p + i_row_string], int(2))
                    #BY contribution to total U mc
                    if i == index_mc_total:
                        value = np.nansum(df_mc["EM_BY_mc_var_normed"].iloc[index_mc_list])
                    else:
                        value = df_mc["EM_BY_mc_var_normed"].iloc[i]
                    apply_number_format(value, ws[col_EM_BY_U_contrib_p + i_row_string], int(2))
                else:
                    apply_style_non_numeric(df_mc['EM_status_BY'].iloc[i], ws[col_EM_BY + i_row_string])
                    apply_style_non_numeric(df_mc['EM_status_BY'].iloc[i], ws[col_EM_BY_U_lower_p + i_row_string])
                    apply_style_non_numeric(df_mc['EM_status_BY'].iloc[i], ws[col_EM_BY_U_upper_p + i_row_string])
                    apply_style_non_numeric(df_mc['EM_status_BY'].iloc[i], ws[col_EM_BY_U_contrib_p + i_row_string])
    
                if df_mc['EM_status_RY'].iloc[i] == "ES":
                    #RY emission as per inventory
                    apply_number_format(df_mc["EM_RY"].iloc[i], ws[col_EM_RY + i_row_string], int(2))
                    #RY U (-) mc
                    apply_number_format(df_mc["EM_RY_mc_U_lower_p"].iloc[i], ws[col_EM_RY_U_lower_p + i_row_string], int(2))
                    #RY U (+) mc
                    apply_number_format(df_mc["EM_RY_mc_U_upper_p"].iloc[i], ws[col_EM_RY_U_upper_p + i_row_string], int(2))
                    #RY contribution to total U mc
                    if i == index_mc_total:
                        value = np.nansum(df_mc["EM_RY_mc_var_normed"].iloc[index_mc_list])
                    else:
                        value = df_mc["EM_RY_mc_var_normed"].iloc[i]
                    apply_number_format(value, ws[col_EM_RY_U_contrib_p + i_row_string], int(2))
                else:
                    apply_style_non_numeric(df_mc['EM_status_RY'].iloc[i], ws[col_EM_RY + i_row_string])
                    apply_style_non_numeric(df_mc['EM_status_RY'].iloc[i], ws[col_EM_RY_U_lower_p + i_row_string])
                    apply_style_non_numeric(df_mc['EM_status_RY'].iloc[i], ws[col_EM_RY_U_upper_p + i_row_string])
                    apply_style_non_numeric(df_mc['EM_status_RY'].iloc[i], ws[col_EM_RY_U_contrib_p + i_row_string])
    
                if df_mc['EM_status_BY'].iloc[i] == "ES" or df_mc['EM_status_RY'].iloc[i] == "ES":
                    apply_number_format(df_mc["EM_trend_normed"].iloc[i], ws[col_trend + i_row_string], int(3))
                    apply_number_format(df_mc["EM_trend_normed_mc_U_lower_p"].iloc[i], ws[col_trend_U_lower_p + i_row_string], int(2))
                    apply_number_format(df_mc["EM_trend_normed_mc_U_upper_p"].iloc[i], ws[col_trend_U_upper_p + i_row_string], int(2))
                    if i == index_mc_total:
                        value = np.nansum(df_mc["EM_trend_normed_mc_var_normed"].iloc[index_mc_list])
                    else:
                        value = df_mc["EM_trend_normed_mc_var_normed"].iloc[i]
                    apply_number_format(value, ws[col_trend_U_contrib_p + i_row_string], int(2))
                else: #both are non-numeric
                    apply_style_non_numeric(df_mc['EM_status_RY'].iloc[i], ws[col_trend + i_row_string])
                    apply_style_non_numeric(df_mc['EM_status_RY'].iloc[i], ws[col_trend_U_lower_p + i_row_string])
                    apply_style_non_numeric(df_mc['EM_status_RY'].iloc[i], ws[col_trend_U_upper_p + i_row_string])
                    apply_style_non_numeric(df_mc['EM_status_RY'].iloc[i], ws[col_trend_U_contrib_p + i_row_string])
        
                for row in ws.iter_rows(min_col = 1, 
                                        min_row = i_row, 
                                        max_col= col_end + 1, #excel_columns
                                        max_row= i_row):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = openpyxl.styles.Alignment(vertical="center", wrap_text=True)
                        if i == index_mc_total:
                            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="8DB4E2")
        
                i_row += 1
                i_row_string = str(i_row) 


    #=================================================================
    #WRITE RESULTS FOR INVENTORY WITH AND WITHOUT LULUCF
    #=================================================================

    if routine == const.ROUTINE_NID:
        #=================================================================
        #WRITE RESULTS FOR INVENTORY WITH AND WITHOUT LULUCF
        #=================================================================
        ws = wb.create_sheet("results_pr_mc_inv_LULUCF")                   
        category_text = "Source category"
        col_cat_name = "proc_code_name"
        index_mc_list = index_mc_inv_with_without_lulucf
        index_mc_list_with_total = index_mc_inv_with_without_lulucf
        index_pr_list_with_total = index_pr_inv_with_without_lulucf
        #print("index_mc_list compound ")
        #print(index_mc_list)

        col_cat = "A" #category: process or compound
        col_EM_BY = "B"
        col_EM_BY_U_lower_p = "C" #in percent of the mean
        col_EM_BY_U_upper_p = "D"
        col_EM_BY_U_mean_p = "E"
        #col_em_u_mean = "E"
        col_EM_BY_end = col_EM_BY_U_mean_p
        col_EM_RY = "F"
        col_EM_RY_U_lower_p = "G" 
        col_EM_RY_U_upper_p = "H"
        col_EM_RY_U_mean_p = "I"
        col_EM_RY_end = col_EM_RY_U_mean_p
        col_trend = "J"
        col_trend_U_lower_p = "K"
        col_trend_U_upper_p = "L"
        col_trend_U_mean_p = "M"
        col_trend_end = col_trend_U_mean_p
        col_end_str = col_trend_end
        
        col_end = [ii for ii in range(len(excel_columns)) if excel_columns[ii] == col_trend_end]
        col_end = col_end[0]
        
        ws.column_dimensions[col_cat].width = 30
        
        
        #******************************HEADER******************************   
        i_row = 1
        i_row_string = str(i_row)
        header_row_start = i_row
        #ws.row_dimensions[i_row].height = 45
    
        #assign data to cells
        ws[col_cat + i_row_string] = category_text
        ws.merge_cells(col_cat + i_row_string + ":" + col_cat + str(i_row + 1) )
    
        ws[col_EM_BY + i_row_string]= "Emissions {}".format(BY_string)
        ws.merge_cells(col_EM_BY + i_row_string + ":" + col_EM_BY_end + i_row_string)
    
        ws[col_EM_RY + i_row_string]= "Emissions {}".format(RY_string)
        ws.merge_cells(col_EM_RY + i_row_string + ":" + col_EM_RY_end + i_row_string)
    
        ws[col_trend + i_row_string]= "Trend {}-{}".format(BY_string, RY_string)
        ws.merge_cells(col_trend + i_row_string + ":" + col_trend_end + i_row_string)
        
        i_row += 1
        i_row_string = str(i_row)
        
        ws[col_EM_BY + i_row_string] = "Value " + unit_string
        ws[col_EM_BY_U_lower_p + i_row_string] = "U(-)%"
        ws[col_EM_BY_U_upper_p + i_row_string] = "U(+)%"
        ws[col_EM_BY_U_mean_p + i_row_string] = "U mean %"
    
        ws[col_EM_RY + i_row_string] = "Value " + unit_string
        ws[col_EM_RY_U_lower_p + i_row_string] = "U(-)%"
        ws[col_EM_RY_U_upper_p + i_row_string] = "U(+)%"
        ws[col_EM_RY_U_mean_p + i_row_string] = "U mean %"
    
        ws[col_trend + i_row_string] = "Value %"
        ws[col_trend_U_lower_p + i_row_string] = "U(-)%"
        ws[col_trend_U_upper_p + i_row_string] = "U(+)%"
        ws[col_trend_U_mean_p + i_row_string] = "U mean %"
    
    
        for row in ws.iter_rows(min_col = 1, 
                                min_row = header_row_start, 
                                max_col= col_end + 1, 
                                max_row= i_row):
            for cell in row:
                cell.border = medium_border
                cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
    
        i_row += 1
        i_row_string = str(i_row)     
        #******************************END OF HEADER******************************     


        #******************************HEADER FOR PR***************************
        ws[col_cat + i_row_string] = "Uncertainty propagation (approach 1)"
        ws.merge_cells(col_cat + i_row_string + ":" + col_end_str + i_row_string)

        for row in ws.iter_rows(min_col = 1, 
                                min_row = i_row, 
                                max_col= col_end + 1, 
                                max_row= i_row):
            for cell in row:
                cell.border = medium_border
                cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        i_row += 1
        i_row_string = str(i_row)
        #******************************END OF HEADER FOR PR********************

        for i in index_pr_list_with_total:
       
            ws[col_cat + i_row_string] = df_pr[col_cat_name].iloc[i]
            
            if df_pr['EM_status_BY'].iloc[i] == "ES":
                #BY emission as per inventory
                apply_number_format(df_pr["EM_BY"].iloc[i], ws[col_EM_BY + i_row_string], int(2))
                #BY U (-) mc
                apply_number_format(df_pr["EM_BY_pr_U_lower_p"].iloc[i], ws[col_EM_BY_U_lower_p + i_row_string], int(2))
                #BY U (+) mc 
                apply_number_format(df_pr["EM_BY_pr_U_upper_p"].iloc[i], ws[col_EM_BY_U_upper_p + i_row_string], int(2))
                #BY contribution to total U mc
                apply_number_format(df_pr["EM_BY_pr_U_mean_p"].iloc[i], ws[col_EM_BY_U_mean_p + i_row_string], int(2))
            else:
                apply_style_non_numeric(df_pr['EM_status_BY'].iloc[i], ws[col_EM_BY + i_row_string])
                apply_style_non_numeric(df_pr['EM_status_BY'].iloc[i], ws[col_EM_BY_U_lower_p + i_row_string])
                apply_style_non_numeric(df_pr['EM_status_BY'].iloc[i], ws[col_EM_BY_U_upper_p + i_row_string])
                apply_style_non_numeric(df_pr['EM_status_BY'].iloc[i], ws[col_EM_BY_U_mean_p + i_row_string])

            if df_pr['EM_status_RY'].iloc[i] == "ES":
                #RY emission as per inventory
                apply_number_format(df_pr["EM_RY"].iloc[i], ws[col_EM_RY + i_row_string], int(2))
                #RY U (-) mc
                apply_number_format(df_pr["EM_RY_pr_U_lower_p"].iloc[i], ws[col_EM_RY_U_lower_p + i_row_string], int(2))
                #RY U (+) mc
                apply_number_format(df_pr["EM_RY_pr_U_upper_p"].iloc[i], ws[col_EM_RY_U_upper_p + i_row_string], int(2))
                #RY contribution to total U mc
                apply_number_format(df_pr["EM_RY_pr_U_mean_p"].iloc[i], ws[col_EM_RY_U_mean_p + i_row_string], int(2))
            else:
                apply_style_non_numeric(df_pr['EM_status_RY'].iloc[i], ws[col_EM_RY + i_row_string])
                apply_style_non_numeric(df_pr['EM_status_RY'].iloc[i], ws[col_EM_RY_U_lower_p + i_row_string])
                apply_style_non_numeric(df_pr['EM_status_RY'].iloc[i], ws[col_EM_RY_U_upper_p + i_row_string])
                apply_style_non_numeric(df_pr['EM_status_RY'].iloc[i], ws[col_EM_RY_U_mean_p + i_row_string])

            if df_pr['EM_status_BY'].iloc[i] == "ES" or df_pr['EM_status_RY'].iloc[i] == "ES":
                apply_number_format(df_pr["EM_trend_normed"].iloc[i], ws[col_trend + i_row_string], int(3))
                apply_number_format(df_pr["EM_trend_normed_pr_U_lower_p"].iloc[i], ws[col_trend_U_lower_p + i_row_string], int(2))
                apply_number_format(df_pr["EM_trend_normed_pr_U_upper_p"].iloc[i], ws[col_trend_U_upper_p + i_row_string], int(2))
                apply_number_format(df_pr["EM_trend_normed_pr_U_mean_p"].iloc[i], ws[col_trend_U_mean_p + i_row_string], int(2))
            else: #both are non-numeric
                apply_style_non_numeric(df_pr['EM_status_RY'].iloc[i], ws[col_trend + i_row_string])
                apply_style_non_numeric(df_pr['EM_status_RY'].iloc[i], ws[col_trend_U_lower_p + i_row_string])
                apply_style_non_numeric(df_pr['EM_status_RY'].iloc[i], ws[col_trend_U_upper_p + i_row_string])
                apply_style_non_numeric(df_pr['EM_status_RY'].iloc[i], ws[col_trend_U_mean_p + i_row_string])
    
            for row in ws.iter_rows(min_col = 1, 
                                    min_row = i_row, 
                                    max_col= col_end + 1, #excel_columns
                                    max_row= i_row):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = openpyxl.styles.Alignment(vertical="center", wrap_text=True)
                    #if i == index_pr_total:
                    #    cell.fill = openpyxl.styles.PatternFill("solid", fgColor="8DB4E2")
    
            i_row += 1
            i_row_string = str(i_row)  
            

        #******************************HEADER FOR MC***************************
        ws[col_cat + i_row_string] = "Monte Carlo simulations (approach 2)"
        ws.merge_cells(col_cat + i_row_string + ":" + col_end_str + i_row_string)

        for row in ws.iter_rows(min_col = 1, 
                                min_row = i_row, 
                                max_col= col_end + 1, 
                                max_row= i_row):
            for cell in row:
                cell.border = medium_border
                cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        i_row += 1
        i_row_string = str(i_row)
        #******************************END OF HEADER FOR MC********************

        
        for i in index_mc_list_with_total:
       
            ws[col_cat + i_row_string] = df_mc[col_cat_name].iloc[i]
            
            if df_mc['EM_status_BY'].iloc[i] == "ES":
                #BY emission as per inventory
                apply_number_format(df_mc["EM_BY"].iloc[i], ws[col_EM_BY + i_row_string], int(2))
                #BY U (-) mc
                apply_number_format(df_mc["EM_BY_mc_U_lower_p"].iloc[i], ws[col_EM_BY_U_lower_p + i_row_string], int(2))
                #BY U (+) mc 
                apply_number_format(df_mc["EM_BY_mc_U_upper_p"].iloc[i], ws[col_EM_BY_U_upper_p + i_row_string], int(2))
                #BY contribution to total U mc
                apply_number_format(df_mc["EM_BY_mc_U_mean_p"].iloc[i], ws[col_EM_BY_U_mean_p + i_row_string], int(2))
            else:
                apply_style_non_numeric(df_mc['EM_status_BY'].iloc[i], ws[col_EM_BY + i_row_string])
                apply_style_non_numeric(df_mc['EM_status_BY'].iloc[i], ws[col_EM_BY_U_lower_p + i_row_string])
                apply_style_non_numeric(df_mc['EM_status_BY'].iloc[i], ws[col_EM_BY_U_upper_p + i_row_string])
                apply_style_non_numeric(df_mc['EM_status_BY'].iloc[i], ws[col_EM_BY_U_mean_p + i_row_string])

            if df_mc['EM_status_RY'].iloc[i] == "ES":
                #RY emission as per inventory
                apply_number_format(df_mc["EM_RY"].iloc[i], ws[col_EM_RY + i_row_string], int(2))
                #RY U (-) mc
                apply_number_format(df_mc["EM_RY_mc_U_lower_p"].iloc[i], ws[col_EM_RY_U_lower_p + i_row_string], int(2))
                #RY U (+) mc
                apply_number_format(df_mc["EM_RY_mc_U_upper_p"].iloc[i], ws[col_EM_RY_U_upper_p + i_row_string], int(2))
                #RY contribution to total U mc
                apply_number_format(df_mc["EM_RY_mc_U_mean_p"].iloc[i], ws[col_EM_RY_U_mean_p + i_row_string], int(2))
            else:
                apply_style_non_numeric(df_mc['EM_status_RY'].iloc[i], ws[col_EM_RY + i_row_string])
                apply_style_non_numeric(df_mc['EM_status_RY'].iloc[i], ws[col_EM_RY_U_lower_p + i_row_string])
                apply_style_non_numeric(df_mc['EM_status_RY'].iloc[i], ws[col_EM_RY_U_upper_p + i_row_string])
                apply_style_non_numeric(df_mc['EM_status_RY'].iloc[i], ws[col_EM_RY_U_mean_p + i_row_string])

            if df_mc['EM_status_BY'].iloc[i] == "ES" or df_mc['EM_status_RY'].iloc[i] == "ES":
                apply_number_format(df_mc["EM_trend_normed"].iloc[i], ws[col_trend + i_row_string], int(3))
                apply_number_format(df_mc["EM_trend_normed_mc_U_lower_p"].iloc[i], ws[col_trend_U_lower_p + i_row_string], int(2))
                apply_number_format(df_mc["EM_trend_normed_mc_U_upper_p"].iloc[i], ws[col_trend_U_upper_p + i_row_string], int(2))
                apply_number_format(df_mc["EM_trend_normed_mc_U_mean_p"].iloc[i], ws[col_trend_U_mean_p + i_row_string], int(2))
            else: #both are non-numeric
                apply_style_non_numeric(df_mc['EM_status_RY'].iloc[i], ws[col_trend + i_row_string])
                apply_style_non_numeric(df_mc['EM_status_RY'].iloc[i], ws[col_trend_U_lower_p + i_row_string])
                apply_style_non_numeric(df_mc['EM_status_RY'].iloc[i], ws[col_trend_U_upper_p + i_row_string])
                apply_style_non_numeric(df_mc['EM_status_RY'].iloc[i], ws[col_trend_U_mean_p + i_row_string])
    
            for row in ws.iter_rows(min_col = 1, 
                                    min_row = i_row, 
                                    max_col= col_end + 1, #excel_columns
                                    max_row= i_row):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = openpyxl.styles.Alignment(vertical="center", wrap_text=True)
                    #if i == index_mc_total:
                    #    cell.fill = openpyxl.styles.PatternFill("solid", fgColor="8DB4E2")
    
            i_row += 1
            i_row_string = str(i_row)         
            
    wb.save(filename_out)
    return None


#HINT below are various functions to handle style in Excel.

excel_columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", 
                 "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]


#styles for output excel files
thin_border = openpyxl.styles.borders.Border(
        left=openpyxl.styles.borders.Side(style='thin'), 
        right=openpyxl.styles.borders.Side(style='thin'), 
        top=openpyxl.styles.borders.Side(style='thin'), 
        bottom=openpyxl.styles.borders.Side(style='thin'))
#double = openpyxl.styles.borders.Side(border_style="double")
medium_border = openpyxl.styles.borders.Border(
        left=openpyxl.styles.borders.Side(style='medium'), 
        right=openpyxl.styles.borders.Side(style='medium'), 
        top=openpyxl.styles.borders.Side(style='medium'), 
        bottom=openpyxl.styles.borders.Side(style='medium'))

top_bottom_medium_border = openpyxl.styles.borders.Border(
        left=openpyxl.styles.borders.Side(style='thin'), 
        right=openpyxl.styles.borders.Side(style='thin'), 
        top=openpyxl.styles.borders.Side(style='medium'), 
        bottom=openpyxl.styles.borders.Side(style='medium'))

left_right_medium_border = openpyxl.styles.borders.Border(
        left=openpyxl.styles.borders.Side(style='medium'), 
        right=openpyxl.styles.borders.Side(style='medium'), 
        top=openpyxl.styles.borders.Side(style='thin'), 
        bottom=openpyxl.styles.borders.Side(style='thin'))

bottom_medium_border = openpyxl.styles.borders.Border(
        left=openpyxl.styles.borders.Side(style='thin'), 
        right=openpyxl.styles.borders.Side(style='thin'), 
        top=openpyxl.styles.borders.Side(style='thin'), 
        bottom=openpyxl.styles.borders.Side(style='medium'))

top_medium_border = openpyxl.styles.borders.Border(
        left=openpyxl.styles.borders.Side(style='thin'), 
        right=openpyxl.styles.borders.Side(style='thin'), 
        top=openpyxl.styles.borders.Side(style='medium'), 
        bottom=openpyxl.styles.borders.Side(style='thin'))


def apply_style_non_numeric(code, cell):
    
    if code == "NA":
        cell.value = 'NA'
        #cell.style = 'Bad'

    elif code == "NE":
        cell.value = 'NE'
        #cell.style = 'Input'
        
    elif code == "IE":
        cell.value = 'IE'
        #cell.style = 'Neutral'

    elif code == "NO":
        cell.value = 'NO'
        #cell.style = 'Good'

    elif code == "ES":
        cell.value = 'ES'
        #cell.style = 'Good'
        
    #change for sub 2023: no specific color for notation keys in all tables.
    cell.fill = openpyxl.styles.PatternFill("solid", fgColor= const.COLOR_GREY)
    cell.alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center", wrap_text=True) 
    
    return None

def apply_style_dEM(cell):
    cell.value = "d.EM"
    cell.alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center", wrap_text=True) 
    cell.fill = openpyxl.styles.PatternFill("solid", fgColor= const.COLOR_GREY)
    return None

def apply_style_empty(cell):
    #cell.value = ""
    cell.alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center", wrap_text=True) 
    cell.fill = openpyxl.styles.PatternFill("solid", fgColor= const.COLOR_GREY)
    return None

def apply_style_cumul_KCA(val, cell):

    color_set = "FFFFFF" #00FFFFFF
    if val>= float(0.055):
        color_set = "C5D9F1"
        #color_set = openpyxl.styles.colors.Color(theme=6, tint=0.8)#light green
    if val>= float(0.105):
        color_set = "8DB4E2"
        #color_set = openpyxl.styles.colors.Color(theme=6, tint=0.5)
    if val>= float(0.305):
        color_set = "538DD5"
        #color_set = openpyxl.styles.colors.Color(theme=6, tint=0.2)
    
    cell.value = val*float(100.0)
    cell.number_format = const.FORMAT_VAL_0D #'##0'
    cell.border = left_right_medium_border
    cell.fill = openpyxl.styles.PatternFill("solid", fgColor=color_set)
    cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    return None

def get_number_format(value, no_sign):
    
    """
    Make sure the number to be written is written witht eadequate number of digit.
    
    no_sign is the total number of cifer to be seen.
    
    """
    val_format = const.FORMAT_VAL_3D
    #check that val is a number
    if isinstance(value, Number):
        val = abs(value)
        
        if no_sign == int(1):
            if val >= float(1.0):
                val_format = const.FORMAT_VAL_0D
            elif val < float(1.0) and val >= float(0.1):
                val_format = const.FORMAT_VAL_1D
            elif val < float(0.1) and val >= float(0.01):
                val_format = const.FORMAT_VAL_2D
            elif val < float(0.01) and val >= float(0.001):
                val_format = const.FORMAT_VAL_3D
            elif val < float(0.001) and val >= float(0.0001):
                val_format = const.FORMAT_VAL_4D
            elif val < float(0.0001) and val >= float(0.00001):
                val_format = const.FORMAT_VAL_5D
            elif val < float(0.00001):
                val_format = const.FORMAT_VAL_6D
    
        elif no_sign == int(2):
    
            if val >= float(10):
                val_format = const.FORMAT_VAL_0D
            elif val < float(10) and val >= float(1.0):
                val_format = const.FORMAT_VAL_1D
            elif val < float(1.0) and val >= float(0.1):
                val_format = const.FORMAT_VAL_2D
            elif val < float(0.1) and val >= float(0.01):
                val_format = const.FORMAT_VAL_3D
            elif val < float(0.01) and val >= float(0.001):
                val_format = const.FORMAT_VAL_4D
            elif val < float(0.001) and val >= float(0.0001):
                val_format = const.FORMAT_VAL_5D
            elif val < float(0.0001):
                val_format = const.FORMAT_VAL_6D
    
        elif no_sign == int(3):
            if val >= float(100.0):
                val_format = const.FORMAT_VAL_0D
            elif val < float(100.0) and val >= float(10.0):
                val_format = const.FORMAT_VAL_1D
            elif val < float(10.0) and val >= float(1.0):
                val_format = const.FORMAT_VAL_2D
            elif val < float(1.0) and val >= float(0.1):
                val_format = const.FORMAT_VAL_3D
            elif val < float(0.1) and val >= float(0.01):
                val_format = const.FORMAT_VAL_4D
            elif val < float(0.01) and val >= float(0.001):
                val_format = const.FORMAT_VAL_5D
            elif val < float(0.001):
                val_format = const.FORMAT_VAL_6D
    
        elif no_sign == int(4):
            if val >= float(1000.0):
                val_format = const.FORMAT_VAL_0D
            elif val < float(1000.0) and val >= float(100.0):
                val_format = const.FORMAT_VAL_1D
            elif val < float(100.0) and val >= float(10.0):
                val_format = const.FORMAT_VAL_2D
            elif val < float(10.0) and val >= float(1.0):
                val_format = const.FORMAT_VAL_3D
            elif val < float(1.0) and val >= float(0.1):
                val_format = const.FORMAT_VAL_4D
            elif val < float(0.1) and val >= float(0.01):
                val_format = const.FORMAT_VAL_5D
            elif val < float(0.01):
                val_format = const.FORMAT_VAL_6D
    
        elif no_sign > int(4):
            val_format = const.FORMAT_VAL_6D
    
    return val_format   

def get_number_format_aligned(value, no_sign):
    
    """
    Make sure the number to be written is written witht eadequate number of digit.
    
    no_sign is the total number of cifer to be seen.
    Numbers are alligned on the decimal point, which requires rouding.
    https://www.extendoffice.com/documents/excel/1951-excel-align-decimal-points.html
    
    """
    val_format = const.FORMAT_VAL_3D
    #check that val is a number
    if isinstance(value, Number):
        val = abs(value)
        
        if no_sign == int(1):
            if val >= float(1.0):
                val = round(val)
                val_format = const.FORMAT_VAL_ALIGNED_1D
            elif val < float(1.0) and val >= float(0.1):
                val_format = const.FORMAT_VAL_ALIGNED_1D
            elif val < float(0.1) and val >= float(0.01):
                val_format = const.FORMAT_VAL_ALIGNED_2D
            elif val < float(0.01) and val >= float(0.001):
                val_format = const.FORMAT_VAL_ALIGNED_3D
            elif val < float(0.001) and val >= float(0.0001):
                val_format = const.FORMAT_VAL_ALIGNED_4D
            elif val < float(0.0001) and val >= float(0.00001):
                val_format = const.FORMAT_VAL_ALIGNED_5D
            elif val < float(0.00001):
                val_format = const.FORMAT_VAL_ALIGNED_6D
    
        elif no_sign == int(2):    
            if val >= float(10.0):
                val = round(val)
                val_format = const.FORMAT_VAL_ALIGNED_1D
            elif val < float(10.0) and val >= float(1.0):
                val_format = const.FORMAT_VAL_ALIGNED_1D
            elif val < float(1.0) and val >= float(0.1):
                val_format = const.FORMAT_VAL_ALIGNED_2D
            elif val < float(0.1) and val >= float(0.01):
                val_format = const.FORMAT_VAL_ALIGNED_3D
            elif val < float(0.01) and val >= float(0.001):
                val_format = const.FORMAT_VAL_ALIGNED_4D
            elif val < float(0.001) and val >= float(0.0001):
                val_format = const.FORMAT_VAL_ALIGNED_5D
            elif val < float(0.0001):
                val_format = const.FORMAT_VAL_ALIGNED_6D
    
        elif no_sign == int(3):
            if val >= float(100.0):
                val = round(val)
                val_format = const.FORMAT_VAL_ALIGNED_1D
            elif val < float(100.0) and val >= float(10.0):
                val_format = const.FORMAT_VAL_ALIGNED_1D
            elif val < float(10.0) and val >= float(1.0):
                val_format = const.FORMAT_VAL_ALIGNED_2D
            elif val < float(1.0) and val >= float(0.1):
                val_format = const.FORMAT_VAL_ALIGNED_3D
            elif val < float(0.1) and val >= float(0.01):
                val_format = const.FORMAT_VAL_ALIGNED_4D
            elif val < float(0.01) and val >= float(0.001):
                val_format = const.FORMAT_VAL_ALIGNED_5D
            elif val < float(0.001):
                val_format = const.FORMAT_VAL_ALIGNED_6D
    
        elif no_sign == int(4):
            if val >= float(1000.0):
                val = round(val)
                val_format = const.FORMAT_VAL_ALIGNED_1D
            elif val < float(1000.0) and val >= float(100.0):
                val_format = const.FORMAT_VAL_ALIGNED_1D
            elif val < float(100.0) and val >= float(10.0):
                val_format = const.FORMAT_VAL_ALIGNED_2D
            elif val < float(10.0) and val >= float(1.0):
                val_format = const.FORMAT_VAL_ALIGNED_3D
            elif val < float(1.0) and val >= float(0.1):
                val_format = const.FORMAT_VAL_ALIGNED_4D
            elif val < float(0.1) and val >= float(0.01):
                val_format = const.FORMAT_VAL_ALIGNED_5D
            elif val < float(0.01):
                val_format = const.FORMAT_VAL_ALIGNED_6D
    
        elif no_sign > int(4):
            if val >= float(10000.0):
                val = round(val)
            val_format = const.FORMAT_VAL_ALIGNED_6D
    
    return val, val_format    
    
def apply_number_format(value, cell, no_sign):
    
    """
    Make sure the number to be written is written witht eadequate number of digit.
    
    no_sign is the total number of cifer to be seen.
    
    """
    if isinstance(value, Number):
        number_format = get_number_format(value, no_sign)
        #if number_format is not None:          
        cell.value = value
        cell.number_format = number_format
    
    return None

def apply_number_format_aligned(value, cell, no_sign):
    
    """
    Make sure the number to be written is written witht eadequate number of digit.
    
    no_sign is the total number of cifer to be seen.
    
    TODO GMY 20230214: it does not work so far.
    The correct number of digit is shown but numbers are not alligned on the decimal point.
    
    """
    if isinstance(value, Number):
        value, number_format = get_number_format_aligned(value, no_sign)
        #if number_format is not None:          
        cell.value = value
        cell.number_format = number_format
    
    return None
